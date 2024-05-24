import os
import time

from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By


class Samsung_Extraction:
    def __init__(self):
        self.driver = webdriver.Chrome()
        self.excel_filepath = r'E:\Kintan\POC2-210623\Samsung_300623.xlsx'
        self.website_url= "https://www.samsung.com/in"
        self.serial_no = 1

    def run(self):
        self.driver.maximize_window()
        self.driver.get(self.website_url)
        time.sleep(4)
        self.click()
        time.sleep(5)
        self.process()
        time.sleep(4)

    def click(self):
        time.sleep(4)
        # Click on the home appliances menu
        home_app = self.driver.find_element(By.XPATH,
                                            """//button[@class='nv00-gnb__l0-menu-btn' and @role='menuitem' and @an-la='home appliances']""")
        print("element found")
        home_app.click()
        time.sleep(2)
        # Click on the "All Refrigerators" link
        all_refrigerators_link = self.driver.find_element(By.XPATH,
                                              """//li[contains(@class, "nv00-gnb__l2-menu")]/a[normalize-space()="All Refrigerators"]""")
        all_refrigerators_link.click()
        time.sleep(4)
        # Scroll down and click on "View more" multiple times to load all refrigerator models
        scroll_counter = 1
        while scroll_counter < 6:
            try:
                print(scroll_counter)
                view_more = self.driver.find_element(By.XPATH, "//a[normalize-space()='View more']")
                self.driver.execute_script("arguments[0].scrollIntoView(true);", view_more)
                time.sleep(2)
                self.driver.execute_script("arguments[0].click();", view_more)
                time.sleep(4)
                scroll_counter += 1
            except NoSuchElementException:
                break

        time.sleep(4)

    def get_number_of_models(self):
        # Get the number of refrigerator models displayed on the page
        model_elements = self.driver.find_elements(By.XPATH,
                                                   """//*[@id="product-range"]/div/div[2]/div[2]/div[2]/div""")
        print(len(model_elements))
        return len(model_elements)

    def process(self):
        self.driver.execute_script("window.scrollTo(0, 0);")
        num_models = self.get_number_of_models()
        for i in range(1, num_models + 1):
            print(i)
            # Click on the "View more" button of each refrigerator model
            model_view = self.driver.find_element(By.XPATH,
                                                  f"""//*[@id="product-range"]/div/div[2]/div[2]/div[2]/div[{i}]//button/span[contains(text(),'View more')]""")
            self.driver.execute_script("arguments[0].scrollIntoView(true);", model_view)
            self.driver.execute_script("arguments[0].click();", model_view)
            time.sleep(8)
            # Extract data from the model
            title_text = self.extract_title()
            price_text = self.extract_price()
            capacity_text = self.extract_capacity()
            width = self.extract_width()
            depth = self.extract_depth()
            height = self.extract_height()
            weight_text = self.extract_weight()
            rating_text = self.extract_rating()

            # Print and write the extracted data to the Excel file
            print(self.serial_no, '--', title_text, price_text, capacity_text, width, depth, height, weight_text,
                  rating_text)
            particular_lst = [self.serial_no, title_text, price_text, capacity_text, width, depth, height, weight_text,
                              rating_text]
            self.write_to_excel(particular_lst)
            self.serial_no += 1
            # Close the model
            model_close = self.driver.find_element(By.XPATH,
                                                   """//*[@id="content"]/div/div/div[3]/div/div[4]/div/div/div/button""")
            self.driver.execute_script("arguments[0].scrollIntoView(true);", model_close)
            self.driver.execute_script("arguments[0].click();", model_close)

        time.sleep(2)

    def extract_title(self):
        try:
            title = self.driver.find_element(By.XPATH,
                                             """//*[@id="content"]/div/div/div[3]/div/div[4]/div/div/div/div/div/div/div/div[2]/div[1]/div[1]/div/div/p[1]""")
            title_text = title.text.strip() if title.text else None
        except NoSuchElementException:
            title_text = None
        return title_text

    def extract_price(self):
        try:
            price = self.driver.find_element(By.XPATH,
                                                       f"//div[contains(normalize-space(@class), 'quick-view__price-container')]/p[contains(.//span[@class='hidden'], 'Total Price:')]")
            price_text = price.text.split("Total Price:")[1].strip() if price.text else None
        except NoSuchElementException:
            price_text = None
        return price_text

    def extract_capacity(self):
        try:
            capacity = self.driver.find_element(By.XPATH,
                                                "//li[contains(.//p, 'Net Total(Liter)')]/p[@class='quick-view__spec-text']")
            capacity_text = capacity.text.strip() if capacity.text else None
        except NoSuchElementException:
            try:
                capacity = self.driver.find_element(By.XPATH,
                                                    "//li[contains(.//p, 'Gross Total(Liter)')]/p[@class='quick-view__spec-text']")
                capacity_text = capacity.text.strip() if capacity.text else None
            except NoSuchElementException:
                capacity_text = None
        return capacity_text

    def extract_width(self):
        try:
            width = self.driver.find_element(By.XPATH,
                                                "//li[contains(.//p, 'Net Width(mm)')]/p[@class='quick-view__spec-text']")
            width_text = width.text.strip() if width.text else None
        except NoSuchElementException:
            try:
                width = self.driver.find_element(By.XPATH,
                                                 "//li[contains(.//p, 'Net Dimension (WxHxD)(mm)')]/p[@class='quick-view__spec-text']")
                width_text= width.text.strip() if width.text else None
            except NoSuchElementException:
                width_text= None

            if width_text:
                if width_text.__contains__('×'):
                    parts = width_text.split('×')
                    width_text = parts[0].strip()
        return width_text

    def extract_depth(self):
        try:
            depth = self.driver.find_element(By.XPATH,
                                                "//li[contains(.//p, 'Net Depth with Door Handle(mm)')]/p[@class='quick-view__spec-text']")
            depth_text = depth.text.strip() if depth.text else None
        except NoSuchElementException:
            try:
                depth = self.driver.find_element(By.XPATH,
                                                 "//li[contains(.//p, 'Net Depth without Door Handle(mm)')]/p[@class='quick-view__spec-text']")
                depth_text = depth.text.strip() if depth.text else None
            except:
                try:
                    depth = self.driver.find_element(By.XPATH,
                                                     "//li[contains(.//p, 'Net Dimension (WxHxD)(mm)')]/p[@class='quick-view__spec-text']")
                    depth_text = depth.text.strip() if depth.text else None
                except NoSuchElementException:
                    depth_text = None

                if depth_text:
                    if depth_text.__contains__('×'):
                        parts = depth_text.split('×')
                        depth_text = parts[2].strip()
        return depth_text

    def extract_height(self):
        try:
            height = self.driver.find_element(By.XPATH,
                                                "//li[contains(.//p, 'Net Case Height with Hinge(mm)')]/p[@class='quick-view__spec-text']")
            height_text = height.text.strip() if height.text else None
        except NoSuchElementException:
            try:
                height = self.driver.find_element(By.XPATH,
                                                 "//li[contains(.//p, 'Net Dimension (WxHxD)(mm)')]/p[@class='quick-view__spec-text']")
                height_text = height.text.strip() if height.text else None
            except NoSuchElementException:
                height_text = None

            if height_text:
                if height_text.__contains__('×'):
                    parts = height_text.split('×')
                    height_text = parts[1].strip()
        return height_text

    def extract_weight(self):
        try:
            weight = self.driver.find_element(By.XPATH,
                                                "//li[contains(.//p, 'Net Weight(kg)')]/p[@class='quick-view__spec-text']")
            weight_text = weight.text.strip() if weight.text else None
        except NoSuchElementException:
            weight_text = None
        return weight_text

    def extract_rating(self):
        try:
            rating = self.driver.find_element(By.XPATH,
                                                "//li[contains(.//p, 'Energy Star Rating')]/p[@class='quick-view__spec-text']")
            rating_text = rating.text.strip() if rating.text else None
        except NoSuchElementException:
            try:
                rating = self.driver.find_element(By.XPATH,
                                                  "//li[contains(.//p, 'Energy Grade')]/p[@class='quick-view__spec-text']")
                rating_text = rating.text.strip() if rating.text else None
            except:
                rating_text = None
        return rating_text


    def write_to_excel(self, data):
        if os.path.isfile(self.excel_filepath):
            wb = load_workbook(self.excel_filepath)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            headers = ['Serial No', 'Title', 'Price', 'Capacity', 'Width', 'Depth', 'Height', 'Weight', 'Rating']
            ws.append(headers)

        ws.append(data)
        wb.save(self.excel_filepath)

if __name__ == '__main__':
    obj = Samsung_Extraction()
    obj.run()
